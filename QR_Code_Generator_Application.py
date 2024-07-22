from openpyxl import Workbook, load_workbook
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
from PIL import Image
import qrcode
import os

count = 6


class Ui_MainWindow(object):

    def Generate(self):

        self.Create_Qr_With_Logo()
        self.Create_Qr_Without_Logo()

    def Create_Qr_Without_Logo(self):

        # This Is Funtcion For Creating a Qr Without Logo
        QRcode = qrcode.QRCode(
            error_correction=qrcode.constants.ERROR_CORRECT_H
        )

        url = self.Line_Edit_1.text()
        QRcode.add_data(url)
        QRcode.make()
        QRcolor = 'Green'

        QRimg = QRcode.make_image(
            fill_color=QRcolor, back_color="white").convert('RGB')

        QRimg.save('generated_qr_codes/qr_Without_Logo.png')

        self.Label_4.setPixmap(QtGui.QPixmap('generated_qr_codes/qr_Without_Logo.png'))

    def Create_Qr_With_Logo(self):

        # This Is Funtcion For Creating a Qr With Logo
        Logo_link = self.Line_Edit_2.text()

        if Logo_link != "":
            logo = Image.open(Logo_link)

        else:
            logo = Image.open("Icons/Logo.jpg")

        basewidth = 100

        wpercent = (basewidth/float(logo.size[0]))
        hsize = int((float(logo.size[1])*float(wpercent)))
        logo = logo.resize((basewidth, hsize), Image.ANTIALIAS)
        QRcode = qrcode.QRCode(
            error_correction=qrcode.constants.ERROR_CORRECT_H
        )

        url = self.Line_Edit_1.text()
        QRcode.add_data(url)
        QRcode.make()
        QRcolor = 'Green'

        QRimg = QRcode.make_image(
            fill_color=QRcolor, back_color="white").convert('RGB')

        pos = ((QRimg.size[0] - logo.size[0]) // 2,
               (QRimg.size[1] - logo.size[1]) // 2)
        QRimg.paste(logo, pos)

        QRimg.save('generated_qr_codes/qr_With_Logo.png')
        self.Label_5.setPixmap(QtGui.QPixmap('generated_qr_codes/qr_With_Logo.png'))

    def save_img(self):

        check = self.Line_Edit_1.text()
        name = self.Line_Edit_3.text()
        path_of_logo = self.Line_Edit_2.text()

        if path_of_logo == "":
            path = 'Icons/Logo.jpg'

        else:
            path = path_of_logo

        if name == "" or check == "":
            msg = QMessageBox()
            msg.setWindowTitle("Error !")
            msg.setText("Please Provide The Given Details !")
            a = msg.exec()

        else:
            wb = load_workbook("Previous Records.xlsx")
            ws = wb.active

            if self.Radio_Button_1.isChecked():
                check_main = 'generated_qr_codes/'+name+'(0).png'
                os.rename('generated_qr_codes/qr_Without_Logo.png', check_main)
                os.remove("generated_qr_codes/qr_With_Logo.png")

            else:
                check_main = 'generated_qr_codes/'+name+'(1).png'
                os.rename('generated_qr_codes/qr_With_Logo.png', check_main)
                os.remove("generated_qr_codes/qr_Without_Logo.png")

            ws['B5'].value = check
            ws['C5'].value = check_main
            ws['D5'].value = path

            ws.insert_rows(5)
            wb.save("Previous Records.xlsx")

            self.Line_Edit_1.setText("")
            self.Line_Edit_2.setText("")
            self.Line_Edit_3.setText("")
            self.Label_4.setPixmap(QtGui.QPixmap(""))
            self.Label_4.setText("QR Without Logo Will Be Shown Here")
            self.Label_5.setPixmap(QtGui.QPixmap(""))
            self.Label_5.setText("QR With Logo Will Be Shown Here")

            msg = QMessageBox()
            msg.setWindowTitle("Success")
            msg.setText("QR Code Is Successfully Generated !")
            b = msg.exec()

    def View_Previous(self):

        wb = load_workbook("Previous Records.xlsx")
        ws = wb.active

        global count

        link = ws['B'+str(count)].value
        Path_Qr = ws['C'+str(count)].value
        Path_logo = ws['D'+str(count)].value

        if (link == None):
            msg = QMessageBox()
            msg.setWindowTitle("Error !")
            msg.setText("No Previous Data Was Found !")
            c = msg.exec()

        else:

            self.Line_Edit_1.setText(link)
            self.Line_Edit_2.setText(Path_logo)
            self.Line_Edit_3.setText(Path_Qr)
            self.Generate()
            count = count+1

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowModality(QtCore.Qt.NonModal)
        MainWindow.resize(500, 680)
        MainWindow.setMinimumSize(QtCore.QSize(500, 680))
        MainWindow.setMaximumSize(QtCore.QSize(500, 680))
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Icons/Logo.jpg"),
                       QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.centralwidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setSpacing(0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.Main_Body = QtWidgets.QFrame(self.centralwidget)
        self.Main_Body.setStyleSheet("QFrame#Main_Body{\n"
                                     "    background-color: rgb(250, 247, 240);\n"
                                     "}\n"
                                     "QPushButton#Push_Button_1{\n"
                                     "    background-color: rgb(188, 206, 248);\n"
                                     "    border: 2px solid rgb(205, 252, 246);\n"
                                     "    border-radius: 20px;\n"
                                     "}\n"
                                     "QPushButton#Push_Button_2,\n"
                                     "QPushButton#Push_Button_3,\n"
                                     "QPushButton#Push_Button_4{\n"
                                     "    background-color: rgb(188, 206, 248);\n"
                                     "    border: 2px solid rgb(205, 252, 246);\n"
                                     "    border-radius: 17px;\n"
                                     "}\n"
                                     "QLineEdit#Line_Edit_1,\n"
                                     "QLineEdit#Line_Edit_2,\n"
                                     "QLineEdit#Line_Edit_3{\n"
                                     "    border: 3px solid rgb(205, 252, 246);\n"
                                     "    border-radius: 17px;\n"
                                     "    padding-right: 20px;\n"
                                     "    padding-left: 15px;\n"
                                     "    padding-bottom: 2px;\n"
                                     "}\n"
                                     "QLabel#Label_4,\n"
                                     "QLabel#Label_5{\n"
                                     "    border: 2px solid #ccc;\n"
                                     "    border-radius: 15px;\n"
                                     "}\n"
                                     "\n"
                                     "Line#line{\n"
                                     "    background-color: rgb(0, 0, 0);\n"
                                     "}")
        self.Main_Body.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.Main_Body.setFrameShadow(QtWidgets.QFrame.Raised)
        self.Main_Body.setObjectName("Main_Body")
        self.Label_1 = QtWidgets.QLabel(self.Main_Body)
        self.Label_1.setGeometry(QtCore.QRect(125, 10, 250, 50))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(16)
        font.setItalic(True)
        self.Label_1.setFont(font)
        self.Label_1.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_1.setObjectName("Label_1")
        self.Label_4 = QtWidgets.QLabel(self.Main_Body)
        self.Label_4.setGeometry(QtCore.QRect(20, 270, 220, 220))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(14)
        font.setItalic(True)
        self.Label_4.setFont(font)
        self.Label_4.setPixmap(QtGui.QPixmap(
            "../Image Scanner App/pics/1-removebg-preview.png"))
        self.Label_4.setScaledContents(True)
        self.Label_4.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_4.setWordWrap(True)
        self.Label_4.setObjectName("Label_4")
        self.Push_Button_2 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_2.setGeometry(QtCore.QRect(150, 185, 200, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(11)
        font.setItalic(True)
        self.Push_Button_2.setFont(font)
        self.Push_Button_2.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Push_Button_2.setObjectName("Push_Button_2")
        self.Push_Button_2.clicked.connect(self.Generate)
        self.Push_Button_4 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_4.setGeometry(QtCore.QRect(260, 620, 200, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(12)
        font.setItalic(True)
        self.Push_Button_4.setFont(font)
        self.Push_Button_4.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Push_Button_4.setObjectName("Push_Button_4")
        self.Push_Button_4.clicked.connect(self.save_img)
        self.Line_Edit_1 = QtWidgets.QLineEdit(self.Main_Body)
        self.Line_Edit_1.setGeometry(QtCore.QRect(120, 80, 350, 35))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.Line_Edit_1.setFont(font)
        self.Line_Edit_1.setInputMask("")
        self.Line_Edit_1.setObjectName("Line_Edit_1")
        self.line = QtWidgets.QFrame(self.Main_Body)
        self.line.setGeometry(QtCore.QRect(0, 240, 500, 2))
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.Label_6 = QtWidgets.QLabel(self.Main_Body)
        self.Label_6.setGeometry(QtCore.QRect(20, 557, 80, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(14)
        font.setItalic(True)
        self.Label_6.setFont(font)
        self.Label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_6.setObjectName("Label_6")
        self.Line_Edit_3 = QtWidgets.QLineEdit(self.Main_Body)
        self.Line_Edit_3.setGeometry(QtCore.QRect(120, 560, 350, 35))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.Line_Edit_3.setFont(font)
        self.Line_Edit_3.setText("")
        self.Line_Edit_3.setObjectName("Line_Edit_3")
        self.Label_2 = QtWidgets.QLabel(self.Main_Body)
        self.Label_2.setGeometry(QtCore.QRect(20, 77, 80, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(14)
        font.setItalic(True)
        self.Label_2.setFont(font)
        self.Label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_2.setObjectName("Label_2")
        self.Label_3 = QtWidgets.QLabel(self.Main_Body)
        self.Label_3.setGeometry(QtCore.QRect(20, 127, 80, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(14)
        font.setItalic(True)
        self.Label_3.setFont(font)
        self.Label_3.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_3.setObjectName("Label_3")
        self.Line_Edit_2 = QtWidgets.QLineEdit(self.Main_Body)
        self.Line_Edit_2.setGeometry(QtCore.QRect(120, 130, 300, 35))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.Line_Edit_2.setFont(font)
        self.Line_Edit_2.setObjectName("Line_Edit_2")
        self.Label_5 = QtWidgets.QLabel(self.Main_Body)
        self.Label_5.setGeometry(QtCore.QRect(260, 270, 220, 220))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(14)
        font.setItalic(True)
        self.Label_5.setFont(font)
        self.Label_5.setPixmap(QtGui.QPixmap(
            "../Image Scanner App/pics/1-removebg-preview.png"))
        self.Label_5.setScaledContents(True)
        self.Label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.Label_5.setWordWrap(True)
        self.Label_5.setObjectName("Label_5")
        self.Radio_Button_1 = QtWidgets.QRadioButton(self.Main_Body)
        self.Radio_Button_1.setGeometry(QtCore.QRect(90, 500, 80, 30))
        self.Radio_Button_1.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Radio_Button_1.setObjectName("Radio_Button_1")
        self.Radio_Button_2 = QtWidgets.QRadioButton(self.Main_Body)
        self.Radio_Button_2.setGeometry(QtCore.QRect(300, 500, 140, 30))
        self.Radio_Button_2.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Radio_Button_2.setObjectName("Radio_Button_2")
        self.Push_Button_1 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_1.setGeometry(QtCore.QRect(430, 127, 40, 40))
        self.Push_Button_1.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.Push_Button_1.setText("")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("Icons/search.png"),
                        QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Push_Button_1.setIcon(icon1)
        self.Push_Button_1.setIconSize(QtCore.QSize(15, 15))
        self.Push_Button_1.setObjectName("Push_Button_1")
        self.Push_Button_3 = QtWidgets.QPushButton(self.Main_Body)
        self.Push_Button_3.setGeometry(QtCore.QRect(40, 620, 200, 35))
        font = QtGui.QFont()
        font.setFamily("Monotype Corsiva")
        font.setPointSize(11)
        font.setItalic(True)
        self.Push_Button_3.setFont(font)
        self.Push_Button_3.setCursor(
            QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap("Icons/left-arrow.png"),
                        QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.Push_Button_3.setIcon(icon2)
        self.Push_Button_3.setIconSize(QtCore.QSize(15, 15))
        self.Push_Button_3.setObjectName("Push_Button_3")
        self.Push_Button_3.clicked.connect(self.View_Previous)
        self.Label_4.raise_()
        self.Label_1.raise_()
        self.Push_Button_2.raise_()
        self.Push_Button_4.raise_()
        self.Line_Edit_1.raise_()
        self.line.raise_()
        self.Label_6.raise_()
        self.Line_Edit_3.raise_()
        self.Label_2.raise_()
        self.Label_3.raise_()
        self.Line_Edit_2.raise_()
        self.Label_5.raise_()
        self.Radio_Button_1.raise_()
        self.Radio_Button_2.raise_()
        self.Push_Button_1.raise_()
        self.Push_Button_3.raise_()
        self.verticalLayout.addWidget(self.Main_Body)
        MainWindow.setCentralWidget(self.centralwidget)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate(
            "MainWindow", "QR_Code_Generator_Application"))
        self.Label_1.setText(_translate("MainWindow", "QR Code Generator App"))
        self.Label_4.setText(_translate(
            "MainWindow", "QR Without Logo Will Be Shown Here"))
        self.Push_Button_2.setText(_translate("MainWindow", "Generate"))
        self.Push_Button_4.setText(_translate("MainWindow", "Submit"))
        self.Line_Edit_1.setPlaceholderText(_translate(
            "MainWindow", "Paste The Link You Want The QR Of"))
        self.Label_6.setText(_translate("MainWindow", "Name :-"))
        self.Line_Edit_3.setPlaceholderText(_translate(
            "MainWindow", "Name You Want To save The QR File"))
        self.Label_2.setText(_translate("MainWindow", "URL :-"))
        self.Label_3.setText(_translate("MainWindow", "lOGO :-"))
        self.Line_Edit_2.setPlaceholderText(
            _translate("MainWindow", "Path of Logo Image"))
        self.Label_5.setText(_translate(
            "MainWindow", "QR With Logo Will Be Shown Here"))
        self.Radio_Button_1.setText(_translate("MainWindow", "Option 1"))
        self.Radio_Button_2.setText(_translate(
            "MainWindow", "Option 2 With Logo"))
        self.Push_Button_3.setText(_translate("MainWindow", "Previous"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())
